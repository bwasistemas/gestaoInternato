from fastapi import FastAPI, Request, HTTPException, Depends, Form
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from app.controllers import router
from app import services
from app.models import Aluno, Especialidade, Local, AgendaCreate, LoginRequest
from datetime import date, timedelta
import json
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

app = FastAPI(title="Internato Manager", version="2.0")

# Configuração de templates e arquivos estáticos
templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")

# Sessões (em produção, usar Redis ou similar)
sessions = {}

# Middleware para verificar autenticação
def require_auth(request: Request):
    session_id = request.cookies.get("session_id")
    if not session_id or session_id not in sessions:
        raise HTTPException(status_code=401, detail="Não autorizado")
    return sessions[session_id]

def get_current_user(request: Request):
    session_id = request.cookies.get("session_id")
    if session_id and session_id in sessions:
        return sessions[session_id]
    return None

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    user = get_current_user(request)
    if user:
        if user.get("tipo") == "aluno":
            return RedirectResponse(url="/aluno-dashboard")
        else:
            return RedirectResponse(url="/dashboard")
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(request: Request, email: str = Form(...), data_nascimento: str = Form(...)):
    user = services.authenticate_user(email, data_nascimento)
    if user:
        import secrets
        session_id = secrets.token_urlsafe(32)
        sessions[session_id] = user

        response = RedirectResponse(url="/dashboard", status_code=302)
        if user.get("tipo") == "aluno":
            response = RedirectResponse(url="/aluno-dashboard", status_code=302)
        else:
            response = RedirectResponse(url="/dashboard", status_code=302)

        response.set_cookie(key="session_id", value=session_id, httponly=True)
        return response
    else:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "E-mail ou data de nascimento inválidos"
        })

@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/", status_code=302)
    response.delete_cookie(key="session_id")
    return response

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")

    dashboard_data = services.get_dashboard_data()
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "dashboard": dashboard_data,
        "user": user
    })

@app.get("/aluno-dashboard", response_class=HTMLResponse)
async def aluno_dashboard(request: Request):
    user = require_auth(request)
    if user.get("tipo") != "aluno":
        return RedirectResponse(url="/dashboard")

    dashboard_data = services.get_aluno_dashboard_data(user["id"])
    if not dashboard_data:
        raise HTTPException(status_code=404, detail="Aluno não encontrado")

    return templates.TemplateResponse("aluno_dashboard.html", {
        "request": request,
        "dashboard": dashboard_data,
        "user": user,
        "today": date.today()
    })

@app.get("/minhas-agendas", response_class=HTMLResponse)
async def minhas_agendas(request: Request):
    user = require_auth(request)
    if user.get("tipo") != "aluno":
        return RedirectResponse(url="/dashboard")

    agendas = services.get_agendas()
    agendas_aluno = [a for a in agendas if a.aluno_id == user["id"]]
    agendas_aluno.sort(key=lambda x: x.data, reverse=True)

    return templates.TemplateResponse("minhas_agendas.html", {
        "request": request,
        "agendas": agendas_aluno,
        "user": user,
        "today": date.today()
    })

@app.get("/alunos", response_class=HTMLResponse)
async def alunos_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    alunos = services.get_alunos()
    especialidades = services.get_especialidades()
    
    # Calcular estatísticas para cada aluno
    agendas = services.get_agendas()
    hoje = date.today()
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    fim_semana = inicio_semana + timedelta(days=6)
    
    for aluno in alunos:
        # Contar agendas do aluno
        agendas_aluno = [a for a in agendas if a.aluno_id == aluno.id]
        aluno.agendas_count = len(agendas_aluno)
        
        # Contar carimbos da semana
        agendas_semana = [a for a in agendas_aluno if inicio_semana <= a.data <= fim_semana]
        aluno.carimbos_semana = sum(1 for a in agendas_semana if a.carimbo)
    
    return templates.TemplateResponse("alunos.html", {
        "request": request,
        "alunos": alunos,
        "especialidades": especialidades,
        "today": date.today().isoformat(),
        "user": user
    })

@app.get("/especialidades", response_class=HTMLResponse)
async def especialidades_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    especialidades = services.get_especialidades()
    return templates.TemplateResponse("especialidades.html", {
        "request": request,
        "especialidades": especialidades,
        "user": user
    })

@app.get("/locais", response_class=HTMLResponse)
async def locais_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    locais = services.get_locais()
    return templates.TemplateResponse("locais.html", {
        "request": request,
        "locais": locais,
        "user": user
    })

@app.get("/locais-estudo", response_class=HTMLResponse)
async def locais_estudo_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    especialidades = services.get_especialidades_completas()
    return templates.TemplateResponse("locais_estudo.html", {
        "request": request,
        "especialidades": especialidades,
        "user": user
    })

@app.get("/agendas", response_class=HTMLResponse)
async def agendas_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    agendas = services.get_agendas()
    alunos = services.get_alunos()
    especialidades = services.get_especialidades_completas()
    return templates.TemplateResponse("agendas.html", {
        "request": request,
        "agendas": agendas,
        "alunos": alunos,
        "especialidades": especialidades,
        "user": user,
        "today": date.today()
    })

@app.get("/distribuicao-carga", response_class=HTMLResponse)
async def distribuicao_carga_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    
    dados_distribuicao = services.get_distribuicao_carga_horaria()
    return templates.TemplateResponse("distribuicao_carga.html", {
        "request": request,
        "user": user,
        **dados_distribuicao
    })

@app.get("/agenda-semanal", response_class=HTMLResponse)
async def agenda_semanal_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    
    especialidades = services.get_especialidades_completas()
    return templates.TemplateResponse("agenda_semanal.html", {
        "request": request,
        "user": user,
        "especialidades": especialidades
    })

@app.get("/presenca", response_class=HTMLResponse)
async def presenca_page(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        return RedirectResponse(url="/aluno-dashboard")
    return templates.TemplateResponse("presenca.html", {
        "request": request,
        "user": user
    })

@app.get("/perfil", response_class=HTMLResponse)
async def perfil_page(request: Request):
    user = require_auth(request)
    # Buscar dados do usuário baseado no tipo
    if user.get("tipo") == "aluno":
        current_user = services.get_aluno_by_id(user["id"])
    else:
        current_user = services.get_aluno_by_email(user["email"])

    return templates.TemplateResponse("perfil.html", {
        "request": request,
        "user": user,
        "aluno": current_user
    })

# API para buscar colegas
@app.get("/api/colegas/{aluno_id}/{data}/{especialidade_id}/{local_id}")
async def get_colegas(aluno_id: int, data: str, especialidade_id: int, local_id: int, request: Request):
    user = require_auth(request)
    if user.get("tipo") != "aluno" or user["id"] != aluno_id:
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    colegas = services.get_colegas_agenda(aluno_id, data, especialidade_id, local_id)
    return colegas

# Incluir rotas da API
app.include_router(router, prefix="/api")

# Endpoints para adicionar/remover dados (simples)
@app.post("/add/{tipo}")
async def add_item(tipo: str, request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    data = await request.json()
    
    if tipo == "aluno":
        services.add_aluno(data)
    elif tipo == "especialidade":
        services.add_especialidade(data)
    elif tipo == "local":
        services.add_local(data)
    elif tipo == "agenda":
        services.add_agenda(data)
    
    return {"message": f"{tipo.capitalize()} adicionado com sucesso"}

@app.delete("/delete/{tipo}/{id}")
async def delete_item(tipo: str, id: int, request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if tipo == "aluno":
        if services.delete_aluno(id):
            return {"message": "Aluno excluído com sucesso"}
        raise HTTPException(status_code=404, detail="Aluno não encontrado")
    
    return {"message": f"{tipo.capitalize()} excluído com sucesso"}

# Rotas de exportação XLS
@app.get("/exportar/agendas-aluno/{aluno_id}")
async def exportar_agendas_aluno_xls(aluno_id: int, request: Request):
    user = require_auth(request)
    if user.get("tipo") != "aluno" or user["id"] != aluno_id:
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Buscar agendas do aluno
    agendas = services.get_agenda_por_aluno(aluno_id, date(2025, 1, 1), date(2025, 12, 31))
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Minhas Agendas"
    
    # Cabeçalhos
    headers = ['Data', 'Dia da Semana', 'Turno', 'Especialidade', 'Local', 'Responsável', 'Status', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="0369a1", end_color="0369a1", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    for row, agenda in enumerate(agendas, 2):
        ws.cell(row=row, column=1, value=agenda.data.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=agenda.data.strftime('%A'))
        ws.cell(row=row, column=3, value=agenda.turno)
        ws.cell(row=row, column=4, value=agenda.especialidade)
        ws.cell(row=row, column=5, value=agenda.local_estudo_nome)
        ws.cell(row=row, column=6, value=agenda.responsavel)
        
        # Status
        if agenda.presenca:
            status = "Presente"
        elif agenda.carimbo:
            status = "Carimbado"
        else:
            status = "Pendente"
        ws.cell(row=row, column=7, value=status)
        
        ws.cell(row=row, column=8, value=agenda.observacoes or "")
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=minhas-agendas-{user['nome']}.xlsx"}
    )

@app.get("/exportar/agenda-semanal")
async def exportar_agenda_semanal_xls(request: Request, data_inicio: str = None, data_fim: str = None):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Se não fornecido, usar semana atual
    if not data_inicio or not data_fim:
        data_inicio = "2025-08-04"
        data_fim = "2025-08-10"
    
    # Buscar agendas do período
    agendas = services.get_agendas()
    agendas_periodo = [a for a in agendas if data_inicio <= a.data.isoformat() <= data_fim]
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda Semanal"
    
    # Cabeçalhos
    headers = ['Aluno', 'Data', 'Dia da Semana', 'Turno', 'Especialidade', 'Local', 'Responsável', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="0369a1", end_color="0369a1", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    for row, agenda in enumerate(agendas_periodo, 2):
        ws.cell(row=row, column=1, value=agenda.aluno_nome)
        ws.cell(row=row, column=2, value=agenda.data.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=agenda.data.strftime('%A'))
        ws.cell(row=row, column=4, value=agenda.turno)
        ws.cell(row=row, column=5, value=agenda.especialidade)
        ws.cell(row=row, column=6, value=agenda.local_estudo_nome)
        ws.cell(row=row, column=7, value=agenda.responsavel)
        
        # Status
        if agenda.presenca:
            status = "Presente"
        elif agenda.carimbo:
            status = "Carimbado"
        else:
            status = "Pendente"
        ws.cell(row=row, column=8, value=status)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=agenda-semanal-{data_inicio}-{data_fim}.xlsx"}
    )

@app.get("/exportar/agendas-todas")
async def exportar_agendas_todas_xls(request: Request):
    user = require_auth(request)
    if user.get("tipo") == "aluno":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Buscar todas as agendas
    agendas = services.get_agendas()
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Todas as Agendas"
    
    # Cabeçalhos
    headers = ['Data', 'Dia da Semana', 'Turno', 'Aluno', 'Especialidade', 'Local de Estudo', 'Responsável', 'Status', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="0369a1", end_color="0369a1", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    for row, agenda in enumerate(agendas, 2):
        ws.cell(row=row, column=1, value=agenda.data.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=agenda.data.strftime('%A'))
        ws.cell(row=row, column=3, value=agenda.turno)
        ws.cell(row=row, column=4, value=agenda.aluno_nome)
        ws.cell(row=row, column=5, value=agenda.especialidade)
        ws.cell(row=row, column=6, value=agenda.local_estudo_nome)
        ws.cell(row=row, column=7, value=agenda.responsavel)
        
        # Status
        if agenda.presenca:
            status = "Presente"
        elif agenda.carimbo:
            status = "Carimbado"
        else:
            status = "Pendente"
        ws.cell(row=row, column=8, value=status)
        
        ws.cell(row=row, column=9, value=agenda.observacoes or "")
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        io.BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=todas-as-agendas.xlsx"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)